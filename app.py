#!/usr/bin/env python
from flask import Flask, render_template, redirect, url_for, request, flash, Response,send_file,make_response, session
from flask_login import LoginManager, login_user, logout_user, current_user, login_required
from flask_sqlalchemy import SQLAlchemy
# from werkzeug.security import generate_password_hash, check_password_hash
from flask_security import UserMixin, RoleMixin, roles_accepted, Security, SQLAlchemySessionUserDatastore
from sqlalchemy import create_engine
from io import BytesIO
from urllib.parse import quote
import pandas as pd
from datetime import timedelta
import os
from openpyxl import load_workbook,drawing
import zipfile
from PIL import Image
from dotenv import load_dotenv
from auth_code import generate_captcha

load_dotenv()

app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = os.getenv("DATABASE_URL")
app.config['SECRET_KEY'] = os.getenv("SECRET_KEY")
app.config['SECURITY_PASSWORD_SALT'] = os.getenv("SECURITY_PASSWORD_SALT")
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=4)
app.config['SECURITY_REGISTERABLE'] = True
app.config['SECURITY_SEND_REGISTER_EMAIL'] = False
app.config['SECURITY_UNAUTHENTICATED_VIEW'] = '/mylogin'
app.config['SECURITY_UNAUTHORIZED_VIEW'] = '/mylogin'
app.config['SECURITY_LOGIN_USER_TEMPLATE'] = 'mylogin.html' # 更改default login page
app.config['ALLOWED_EXTENSIONS'] = {'PDF', 'JPG', 'JPEG', 'PNG', 'HEIC', 'AI'}
app.config['MAX_FILE_SIZE'] =  10 * 1024 * 1024  # 10MB 的位元組
app.config['EXPORT_FOLDER'] = 'EXP_FOLDER'
app.config['審核狀態'] = ['尚未審核','核准參賽','備取','審核不通過']
app.config['GAME'] = {'賽事名稱':'2023 第十屆全國 EMBA 籃球邀請賽', 
                      '主辦':'國立成功大學、國立成功大學EMBA校友會、國立成功大學管理學院EMBA籃球社',
                      '承辦':'國立成功大學管理學院EMBA籃球社',
                      '比賽地點':'台南市成功大學中正堂及台南市健美洋行 3F(台南一中體育館)',
                      '比賽開始日期':'2023 年 11 月 25 日(週六)','比賽結束日期':'2023 年 11 月 26 日(週日)',
                      '參賽組別':{'挑戰賽':'上限 20 隊','經典賽':'上限 8 隊'},
                      '開始繳費日期':'2023年7月10日 09:00','截止繳費日期':'2023年7月21日',
                      '報名表(附件一)回傳日期':'2023/9/11 中午 12:00 前',
                      '賽事狀態':'報名中','賽事管理員':'vicfenny@gmail.com'}

db = SQLAlchemy(app)

  

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'mylogin'

engine = create_engine(app.config['SQLALCHEMY_DATABASE_URI'])

# 自定義未授權處理程序
@login_manager.unauthorized_handler
def unauthorized_callback():
    return redirect(url_for('mylogin')) 

# Setup User Model and link with database
roles_users = db.Table('roles_users',
        db.Column('user_id', db.Integer(), db.ForeignKey('user.id')),
        db.Column('role_id', db.Integer(), db.ForeignKey('role.id'))) 

class User(db.Model, UserMixin):
    __tablename__ = 'user'
    id = db.Column(db.Integer, autoincrement=True, primary_key=True)
    email = db.Column(db.String, unique=True)
    password = db.Column(db.String(255), nullable=False, server_default='')
    active = db.Column(db.Boolean())
    name = db.Column(db.String)
    team_id = db.Column(db.String)
    phone = db.Column(db.String)
    mobile = db.Column(db.String)
    # backreferences the user_id from roles_users table
    roles = db.relationship('Role', secondary=roles_users, backref='roled')
 
class Role(db.Model, RoleMixin):
    __tablename__ = 'role'
    id = db.Column(db.Integer(), primary_key=True)
    name = db.Column(db.String(80), unique=True)
    description = db.Column(db.String)
 
# load users, roles for a session
user_datastore = SQLAlchemySessionUserDatastore(db.session, User, Role)
security = Security(app, user_datastore)

# Load User
@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

login_manager.user_loader(load_user)

# Routes
@app.route('/')
def home():
    game_content = app.config['GAME']
    return render_template('home.html',game_content=game_content)

@app.route('/admin')
@login_required
@roles_accepted('admin')
def admin_dashboard():
        return '系統管理者 Dashboard (Requires admin role)'
    
@app.route('/gamemanager')
@login_required
@roles_accepted('admin', 'gamemanager')
def gamemanager_dashboard():
    return '比賽主辦管理者 Dashboard (Requires gamemanager role)'

@app.route('/user')
@login_required
@roles_accepted('admin', 'gamemanager', 'user')
def user_dashboard():
        return '參賽隊伍 Dashboard (Requires user role)'

@app.route('/mylogin', methods=['GET', 'POST'])
def mylogin():
    if current_user.is_authenticated:
        return redirect(request.referrer or url_for('home'))
    if request.method != 'POST':
        img_url, captcha_ans = generate_captcha(5)
        session['img_url'] = img_url
        session['captcha_ans'] = captcha_ans
        # print(session['captcha_ans'])
    if request.method == 'POST':
        email = request.form['email'].strip()
        password = request.form['password'].strip()
        captcha = request.form['captcha'].strip()
        user = User.query.filter_by(email=email).first()
        if user and user.password == password:
            if captcha == session['captcha_ans']:
                login_user(user)
                flash('登入成功！', 'success')
                return redirect(url_for('get_teams', game_id='2023 第十屆全國 EMBA 籃球邀請賽') )
            else:
                img_url, captcha_ans = generate_captcha(5)
                session['img_url'] = img_url
                session['captcha_ans'] = captcha_ans
                flash('驗證碼錯誤！請重試！', 'danger')
        else:
            img_url, captcha_ans = generate_captcha(5)
            session['img_url'] = img_url
            session['captcha_ans'] = captcha_ans            
            flash('email 或 password 錯誤！請重試！', 'danger')  

    return render_template('mylogin.html', img_url=session['img_url'])

@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('登出成功！', 'success')
    return redirect(url_for('home'))

@app.route('/gamerule')
def gamerule():
    return render_template('rule.html')

# query all teams 
@app.route('/get_teams/<game_id>', methods=['GET', 'POST'])
@login_required
@roles_accepted('admin', 'gamemanager', 'user')
def get_teams(game_id):
    if current_user.has_role('admin') or current_user.has_role('gamemanager'):
        condition = ''
    else:    
        # 只有 user 權限時, 才能看到同隊伍的隊員資料
        condition = f" AND contact_pid='{current_user.id}'"
    sql = f'''SELECT team_id, team_name 報名單位, group_id 參賽組別, status 狀態, qualified 是否合格 FROM team 
        WHERE game_id='{game_id}' {condition} order by 報名單位, 參賽組別'''
    df = pd.DataFrame(engine.execute(sql))
    df.index += 1 
    if df.shape[0] >= 1: #該權限內可看到的隊伍, 若至少有一隊已報名, 則顯示隊伍清單
        # 加'報名單位' hyperlink 至 editteam_member 頁面
        df['報名單位'] = df.apply(lambda x: f"<a href={url_for('editteam_member',team_id=(x.team_id))}>{x.報名單位}</a>", axis=1)
        df = df.drop('team_id', axis=1)
        # 將dataframe 的 是否合格 column 內容為false 置換成“否”, true 置換成“是”,
        df['是否合格'] = df['是否合格'].apply(lambda x: '是' if x else '否')
        #df['是否合格'] = df['是否合格'].apply(lambda x: '<select><option value="true" selected>是</option><option value="false">否</option></select>' if x else '<select><option value="true">是</option><option value="false" selected>否</option></select>')
    else: #該權限內沒有看到任一個個報名隊伍, 則顯示'尚無您的報名隊伍'
        new_data = { '報名單位': '尚無您的報名隊伍'}
        df = df.append(new_data, ignore_index=True)
    # 將html table head 文字靠左對齊
    data_html = df.to_html(classes=['table table-border table-striped'], escape=False).replace('<th', '<th style="text-align: left;"')
        
    return render_template('show_teams.html', outgame_id=game_id, outdata=data_html)

# member data database CRUD

# 從 team table(隊伍資料表) 取得隊伍的team id
def get_teamid_fm_pid(ct_pid):
    sql = f'''SELECT team_id FROM team WHERE contact_pid={ct_pid} '''
    team_id = engine.execute(sql).fetchall()
    # convert [(1,), (6,), (7,), (8,)] into [1, 6, 7, 8]
    ls_team_id = [tid[0] for tid in team_id ]
    return ls_team_id

@app.route('/editteam_member/<int:team_id>', methods=['GET', 'POST'])
@login_required
@roles_accepted('admin', 'gamemanager', 'user')
def editteam_member(team_id):
    can_edit_team = False
    if current_user.has_role('admin') or current_user.has_role('gamemanager'):
        can_edit_team = True
    else:
        cur_u_team_id = get_teamid_fm_pid(current_user.id) 
        if team_id in cur_u_team_id:
            can_edit_team = True
    if can_edit_team:
        sql = f'''SELECT reg_pid,jersey_number 背號,student_name 姓名,grade "EMBA級別",birthday 出生年月日,pid 身分證字號,
            CASE WHEN islimited IS true THEN '✅' ELSE '' end as 限制球員,
            CASE WHEN isteacher IS true THEN '✅' ELSE '' end as 教職員,
            CASE WHEN st_data IS NOT NULL THEN '_S' ELSE '' end as 大頭照
            FROM registration WHERE team_num={team_id} ORDER BY reg_pid'''
        data = engine.execute(sql)
        column_names = data.keys()
        sql_team = f'''SELECT A.team_id,B.id,team_name 報名單位,group_id 參賽組別,name 聯絡人,phone 電話,mobile "LINE ID",email 電子郵件,
                coach 教練,head_coach 領隊,team_captain 隊長,
                CASE WHEN sign_data IS NOT NULL THEN 'Y' ELSE '' END as 系辦蓋章,
                CASE WHEN logo_data IS NOT NULL THEN 'Y' ELSE '' END as "學校Logo",
                CASE WHEN qualified IS true THEN '是' ELSE '否' END as 是否合格, status 狀態
                FROM team A INNER JOIN "user" B ON B.id=A.contact_pid WHERE A.team_id={team_id} '''
        team_data = engine.execute(sql_team).fetchone()
        team_column_names = team_data.keys()
        approval_statuslist = app.config['審核狀態']
    else:
        flash(f"無權限，請確認！","danger")
        ref = request.referrer
        if ref:
            return redirect(ref)
        else:
            return redirect(url_for('home'))
        
    return render_template('output7.html', outdata=data, outheaders=column_names,outteam=team_data,outteamheader=team_column_names,
                           outstatuslist=approval_statuslist)

#新增隊伍資料
@app.route('/add_team/<gid>/<int:ct_pid>', methods=['GET','POST'])
@login_required
def add_team(gid,ct_pid):
    if request.method == "POST":
        team_name = request.values['in_tname'].strip()
        group_name = request.values['in_group'].strip()
        coach_name = request.values['in_coach'].strip()
        headcoach_name = request.values['in_hcoach'].strip()
        team_captain_name = request.values['in_captain'].strip()

        conn = engine.connect()
        trans = conn.begin()
        try:
            sql1 = f''' INSERT INTO team(team_name,game_id,group_id,contact_pid,coach,head_coach,team_captain,status) 
                VALUES('{team_name}','{gid}','{group_name}',{ct_pid},'{coach_name}','{headcoach_name}',
                '{team_captain_name}','尚未審核')  '''
            conn.execute(sql1)
            sql2 = f''' SELECT MAX(team_id) FROM team '''
            tid = conn.execute(sql2).fetchone()[0] #取得新增的自動編號team_id
            #找出所有 input box 為file type, iboxname存放所有input box name, file 存放檔案資料 
            for iboxname,file in request.files.items():
                if file.filename:
                    file_extension = file.filename.rsplit('.', 1)[1].upper()
                    file_size = len(file.read())
                    file.seek(0)  # 將檔案指標重新移回檔案開頭
            #若HTML表單有選擇檔案, 但副檔名不在許可格式清單中, 則禁止更新, 並拋送Exception訊息
            if file and not(file_extension in app.config['ALLOWED_EXTENSIONS']):
                raise Exception(f"{file_extension}檔案格式不能上傳")
            #若HTML表單有選擇檔案, 但檔案大小超過規定, 則禁止新增, 並拋送Exception訊息
            if file and file_size > app.config['MAX_FILE_SIZE']:
                raise Exception(f"檔案太大,無法上傳,不能>{app.config['MAX_FILE_SIZE']/(1024*1024)}MB")
            
            for iboxname, file in request.files.items():
                # 若HTML表單有選擇檔案, 根據input box name而決定更新資料庫對應檔案欄位, 則更新檔案
                if file.filename:
                    if iboxname.upper() == 'IN_FILE_SIGNDOC':
                        sql2 ="UPDATE team SET sign_filename=%s, sign_data=%s WHERE team_id=%s"
                        conn.execute(sql2, file.filename, file.read(), tid)
                    elif iboxname.upper() == 'IN_FILE_LOGO':
                        sql2_1 ="UPDATE team SET logo_filename=%s, logo_data=%s WHERE team_id=%s"
                        conn.execute(sql2_1, file.filename, file.read(), tid)

            trans.commit()
            flash("新增隊伍資料成功","primary")
        except Exception as e:
            trans.rollback()
            flash(f"新增隊伍資料失敗,{str(e)}","danger")
        finally:
            conn.close()
        sql2 = f''' SELECT MAX(team_id) FROM team '''
        new_teamid = engine.execute(sql2).fetchone()[0] #取得新增的自動編號team_id

    return redirect(url_for('editteam_member', team_id=new_teamid))

#修改隊伍(聯絡人)資料
@app.route('/edit_team/<int:tid>/<int:pid>', methods=['GET','POST'])
@login_required
def edit_team(tid,pid):
    if request.method == "POST":
        ct_name = request.values['in_ctname'].strip()
        phone = request.values['in_phone'].strip()
        mobile = request.values['in_mobile'].strip()
        team_name = request.values['in_tname'].strip()
        group_name = request.values['in_group'].strip()
        coach_name = request.values['in_coach'].strip()
        headcoach_name = request.values['in_hcoach'].strip()
        team_captain_name = request.values['in_captain'].strip()
        if current_user.has_role('admin') or current_user.has_role('gamemanager'):
            approval_status = request.values['in_approval'].strip()
            qualified = request.values['in_valid'].strip()
            sql1_1 = f", status='{approval_status}', qualified={qualified}" 
        else:
            sql1_1 = ""

        conn = engine.connect()
        trans = conn.begin()
        try:
            sql1 = f'''UPDATE "user" SET name='{ct_name}',phone='{phone}',mobile='{mobile}' WHERE id={pid} '''
            conn.execute(sql1)
            sql2 = f'''UPDATE team SET team_name='{team_name}', group_id='{group_name}', coach='{coach_name}',
                head_coach='{headcoach_name}', team_captain='{team_captain_name}' {sql1_1}, update_time='now()' 
                WHERE team_id={tid} '''
            conn.execute(sql2)

            #找出所有 input box 為file type, iboxname存放所有input box name, file 存放檔案資料 
            for iboxname,file in request.files.items():
                if file.filename:
                    file_extension = file.filename.rsplit('.', 1)[1].upper()
                    file_size = len(file.read())
                    file.seek(0)  # 將檔案指標重新移回檔案開頭
            #若HTML表單有選擇檔案, 但副檔名不在許可格式清單中, 則禁止更新, 並拋送Exception訊息
            if file and not(file_extension in app.config['ALLOWED_EXTENSIONS']):
                raise Exception(f"{file_extension}檔案格式不能上傳")
            #若HTML表單有選擇檔案, 但檔案大小超過規定, 則禁止新增, 並拋送Exception訊息
            if file and file_size > app.config['MAX_FILE_SIZE']:
                raise Exception(f"檔案太大,無法上傳,不能>{app.config['MAX_FILE_SIZE']}Bytes")
            
            for iboxname, file in request.files.items():
                # 若HTML表單有選擇檔案, 根據input box name而決定更新資料庫對應檔案欄位, 則更新檔案
                if file.filename:
                    if iboxname.upper() == 'IN_FILE_SIGNDOC':
                        sql2 ="UPDATE team SET sign_filename=%s, sign_data=%s WHERE team_id=%s"
                        conn.execute(sql2, file.filename, file.read(), tid)
                    elif iboxname.upper() == 'IN_FILE_LOGO':
                        sql2_1 ="UPDATE team SET logo_filename=%s, logo_data=%s WHERE team_id=%s"
                        conn.execute(sql2_1, file.filename, file.read(), tid)
                        
                # 若HTML表單未選擇檔案, 並且勾選移除已上傳檔案, 則清除資料庫檔案
                elif not(file.filename):
                    if iboxname.upper()=='IN_FILE_SIGNDOC' and request.form.get('in_rmsigndocexistfile')=='SIGNDOC':
                        sql3 ="UPDATE team SET sign_filename=null, sign_data=null WHERE team_id=%s"
                        conn.execute(sql3, tid)
                    elif iboxname.upper()=='IN_FILE_LOGO' and request.form.get('in_rmlogoexistfile')=='LOGO':
                        sql3_1 ="UPDATE team SET logo_filename=null, logo_data=null WHERE team_id=%s"
                        conn.execute(sql3_1, tid)

            trans.commit()
            flash("修改隊伍資料成功","primary")
        except Exception as e:
            trans.rollback()
            flash(f"修改隊伍資料失敗,{str(e)}","danger")
        finally:
            conn.close()
        
    return redirect(url_for('editteam_member', team_id=tid))


# 新增人員
@app.route('/add_person/<int:tid>', methods=['GET','POST'])
@login_required
def add_person(tid):
    if request.method == "POST":
        jersey_number = request.values['in_jersey'].strip() #去除頭尾空白字元
        name = request.values['in_name'].strip()
        grade = request.values['in_grade'].strip()
        birthday = request.values['in_birthday'].strip()
        pid = request.values['in_pid'].strip()
        islimited = request.values['in_limited'].strip()
        isteacher = request.values['in_teacher'].strip()

        conn = engine.connect()
        trans = conn.begin()
        try:
            #找出所有 input box 為file type, iboxname存放所有input box name, file 存放檔案資料 
            for iboxname,file in request.files.items():
                if file.filename:
                    file_extension = file.filename.rsplit('.', 1)[1].upper()
                    file_size = len(file.read())
                    file.seek(0)  # 將檔案指標重新移回檔案開頭
                #若HTML表單有選擇檔案, 但副檔名不在許可格式清單中, 則禁止新增, 並拋送Exception訊息
                if file and not(file_extension in app.config['ALLOWED_EXTENSIONS']):
                    raise Exception(f"{file_extension}檔案格式不能上傳")
                #若HTML表單有選擇檔案, 但檔案大小超過規定, 則禁止新增, 並拋送Exception訊息
                if file and file_size > app.config['MAX_FILE_SIZE']:
                    raise Exception(f"檔案太大,無法上傳,不能>{app.config['MAX_FILE_SIZE']}Bytes")
            #team_num = get_teamid_fm_pid(current_user.id)
            sql1 = f'''insert into registration(jersey_number,student_name,grade,birthday,pid,islimited,isteacher,team_num ) 
                values ('{jersey_number}','{name}','{grade}','{birthday}','{pid}','{islimited}','{isteacher}',{tid})'''
            conn.execute(sql1)
            sql2 = f'''select max(reg_pid) from registration'''
            result = conn.execute(sql2)
            new_reg_pid = result.fetchone()[0] #取得新增人員對應的自動編號id(reg_id)
            for iboxname, file in request.files.items():
                # 若HTML表單有選擇檔案, 根據input box name而決定更新資料庫對應檔案欄位, 則新增檔案
                if file.filename:
                    if iboxname.upper() == 'IN_FILE_STUID':
                        sql2 ="UPDATE registration SET st_filename=%s, st_data=%s WHERE reg_pid=%s"
                        '''
                        case 'IN_FILE_PID':
                            sql2 ="UPDATE registration SET pid_filename=%s, pid_data=%s WHERE pid=%s"
                        case 'IN_FILE_STUID':
                            
                        case 'IN_FILE_ENROLL':
                            sql2 ="UPDATE registration SET er_filename=%s, er_data=%s WHERE pid=%s"
                        '''
                    conn.execute(sql2, file.filename, file.read(), new_reg_pid) 
            trans.commit()
            flash("新增人員成功","primary")
        except Exception as e:
            trans.rollback()
            flash(f"新增人員失敗,{str(e)}","danger")
        finally:
            conn.close()
        #print(school+name+email)
    return redirect(url_for('editteam_member', team_id=tid))

#修改人員資料
@app.route('/edit_person/<int:reg_pid>', methods=['GET','POST'])
@login_required
def edit_person(reg_pid):
    if request.method == "POST":
        team_num = request.values['in_teamid']
        jersey_number = request.values['in_jersey'].strip()
        name = request.values['in_name'].strip()
        grade = request.values['in_grade'].strip()
        birthday = request.values['in_birthday'].strip()
        pid = request.values['in_pid'].strip()
        islimited = request.values['in_limited'].strip()
        isteacher = request.values['in_teacher'].strip()
        
        conn = engine.connect()
        trans = conn.begin()
        try:
            #找出所有 input box 為file type, iboxname存放所有input box name, file 存放檔案資料 
            for iboxname,file in request.files.items():
                if file.filename:
                    file_extension = file.filename.rsplit('.', 1)[1].upper()
                    file_size = len(file.read())
                    file.seek(0)  # 將檔案指標重新移回檔案開頭
                #若HTML表單有選擇檔案, 但副檔名不在許可格式清單中, 則禁止更新, 並拋送Exception訊息
                if file and not(file_extension in app.config['ALLOWED_EXTENSIONS']):
                    raise Exception(f"{file_extension}檔案格式不能上傳")
                #若HTML表單有選擇檔案, 但檔案大小超過規定, 則禁止新增, 並拋送Exception訊息
                if file and file_size > app.config['MAX_FILE_SIZE']:
                    raise Exception(f"檔案太大,無法上傳,不能>{app.config['MAX_FILE_SIZE']}Bytes")
            sql1 = f'''UPDATE registration SET jersey_number='{jersey_number}',student_name='{name}',
                birthday='{birthday}',pid='{pid}',grade='{grade}',islimited={islimited}, isteacher={isteacher},
                update_time='now()' WHERE reg_pid={reg_pid} '''
            conn.execute(sql1)

            for iboxname, file in request.files.items():
                # 若HTML表單有選擇檔案, 根據input box name而決定更新資料庫對應檔案欄位, 則更新檔案
                if file.filename:
                    if iboxname.upper() == 'IN_FILE_STUID':
                        sql2 ="UPDATE registration SET st_filename=%s, st_data=%s WHERE reg_pid=%s"
                        conn.execute(sql2, file.filename, file.read(), reg_pid)
                        '''
                        case 'IN_FILE_PID':
                            sql2 ="UPDATE registration SET pid_filename=%s, pid_data=%s WHERE pid=%s"
                            conn.execute(sql2, file.filename, file.read(), pid)       
                        case 'IN_FILE_ENROLL':
                            sql2 ="UPDATE registration SET er_filename=%s, er_data=%s WHERE pid=%s"
                            conn.execute(sql2, file.filename, file.read(), pid)    
                        '''
                # 若HTML表單未選擇檔案, 並且勾選移除已上傳檔案, 則清除資料庫檔案
                elif not(file.filename):
                    if iboxname.upper()=='IN_FILE_STUID' and request.form.get('in_rmstuexistfile')=='STUID':
                        sql3 ="UPDATE registration SET st_filename=null, st_data=null WHERE reg_pid=%s"
                        conn.execute(sql3, reg_pid)
                    '''
                    elif iboxname.upper()=='IN_FILE_PID' and request.form.get('in_rmpidexistfile')=='PID':
                        sql3 ="UPDATE registration SET pid_filename=null, pid_data=null WHERE pid=%s"
                        conn.execute(sql3, pid)
                    elif iboxname.upper()=='IN_FILE_ENROLL' and request.form.get('in_rmenrollexistfile')=='ENROLL':
                        sql3 ="UPDATE registration SET er_filename=null, er_data=null WHERE pid=%s"
                        conn.execute(sql3, pid)    
                    '''
            trans.commit()
            flash("修改人員成功","primary")
        except Exception as e:
            trans.rollback()
            flash(f"修改人員失敗,{str(e)}","danger")
        finally:
            conn.close()
        #team_num = get_teamid_fm_pid(reg_pid)
    return redirect(url_for("editteam_member", team_id=team_num))

#刪除人員資料
@app.route('/del_person/<int:reg_pid>', methods=['GET','POST'])
@login_required
def del_person(reg_pid):
    if request.method == "POST":
        #team_num = get_teamid_fm_pid(reg_pid)
        team_num = request.values['in_teamid']
        conn = engine.connect()
        trans = conn.begin()
        try:   
            sql = f'''DELETE FROM registration WHERE reg_pid={reg_pid} '''
            conn.execute(sql)
            trans.commit()
            flash("刪除人員成功","primary")
        except Exception as e:
            trans.rollback()
            flash(f"刪除人員失敗,{str(e)}","danger")
        finally:
            conn.close()
        
    return redirect(url_for("editteam_member", team_id=team_num))

# create showfile function 
@app.route('/showfile/<ftype>/<reg_pid>', methods=['GET','POST'])
@login_required
def showfile(ftype,reg_pid):
    if request.method == "GET":
        try:
            match ftype:
                case 'PID':
                    fname = 'pid_filename'
                    fdata = 'pid_data'
                case 'STUID':
                    fname = 'st_filename'
                    fdata = 'st_data'
                    tablename = 'registration'
                    idname = 'reg_pid'
                case 'ENROLL':
                    fname = 'er_filename'
                    fdata = 'er_data'
                case 'SIGNDOC':
                    fname = 'sign_filename'
                    fdata = 'sign_data'
                    tablename = 'team'
                    idname = 'team_id'
                case 'LOGO':
                    fname = 'logo_filename'
                    fdata = 'logo_data'
                    tablename = 'team'
                    idname = 'team_id'
            sql = f"select {fname},{fdata} from {tablename} where {idname}={reg_pid}"
            file_data = engine.execute(sql).fetchone()
            file_stream = BytesIO(file_data[fdata])
            file_name = file_data[fname]
            file_extension = file_name.split(".")[-1].upper()
            headers = {}
            #為了傳檔案類型給Browser, 才能正確顯示檔案
            match file_extension:
                case 'JPG' | 'JPEG':
                    filetype = 'image/jpeg'
                case 'PNG':
                    filetype = 'image/png'
                case 'PDF':
                    filetype = 'application/pdf'
                case 'HEIC':
                    filetype = 'image/heic'
                case 'AI':
                    filetype = "application/illustrator"
                case _:
                    filetype = 'na'
            #flash("Show檔案成功","primary")
            headers = {	'Content-Type': filetype,
                        'Content-Disposition': f'''inline; filename="{quote(file_data[fname])}"'''}
    
        except Exception as e:
            flash(f"Show檔案失敗,{str(e)}","danger")
    # 以下 send_file->download_name 需搭配python v3.11
    #return send_file(file_stream, mimetype=filetype, download_name=(quote(file_data['pid_filename'])), as_attachment=False)
    return Response(file_stream, headers=headers)
 
@app.route('/download/<int:team_id>', methods=['GET'])
@login_required
@roles_accepted('admin', 'gamemanager', 'user')
def download(team_id):

    sql_team = f'''SELECT A.game_id,A.team_id,B.id,team_name 報名單位,group_id 參賽組別,name 聯絡人,phone 電話,mobile "LINE_ID",email 電子郵件,
            coach 教練,head_coach 領隊,team_captain 隊長
            FROM team A INNER JOIN "user" B ON B.id=A.contact_pid WHERE A.team_id={team_id} '''
    df1 = pd.DataFrame(engine.execute(sql_team))

    sql_reg = f'''SELECT jersey_number 背號,student_name 姓名,grade "EMBA級別",birthday 出生年月日,pid 身分證字號,
        CASE WHEN islimited IS true THEN 'ｖ' ELSE '' end as 限制球員,
        CASE WHEN isteacher IS true THEN 'ｖ' ELSE '' end as 教職員
        from REGISTRATION WHERE team_num={team_id} ORDER BY reg_pid'''
    df2 = pd.DataFrame(engine.execute(sql_reg))
    
    sql_reg_photo = f'''SELECT team_num, student_name 姓名, st_data FROM registration  
                    WHERE team_num={team_id} ORDER BY reg_pid'''
    data = engine.execute(sql_reg_photo)

    # 載入報名表範本
    template_path = os.path.join(app.root_path, 'templates', '報名表.xlsx')
    workbook = load_workbook(template_path)
    sheet = workbook['報名表資料']
    # 將 賽事名稱 資料存放至指定儲存格
    sheet['C2'].value = df1['game_id'].values[0]+'報名表(資料)'
    # 將 報名單位 資料存放至指定儲存格
    sheet['C3'].value = df1['報名單位'].values[0]
    # 將 聯絡人 資料存放至指定儲存格
    sheet['F3'].value = df1['聯絡人'].values[0]
    # 將 電話 資料存放至指定儲存格
    sheet['C4'].value = df1['電話'].values[0]
    # 將 LINE_ID 資料存放至指定儲存格
    sheet['F4'].value = df1['LINE_ID'].values[0]
    # 將 電子郵件 資料從存放至指定儲存格
    sheet['C5'].value = df1['電子郵件'].values[0]
    # 將 參賽組別 資料從存放至指定儲存格
    sheet['C6'].value = df1['參賽組別'].values[0]
    # 將 教練 資料從存放至指定儲存格
    sheet['C7'].value = df1['教練'].values[0]
    # 將 領隊 資料從存放至指定儲存格
    sheet['E7'].value = df1['領隊'].values[0]
    # 將 隊長 資料從存放至指定儲存格
    sheet['G7'].value = df1['隊長'].values[0]

    # 將 DataFrame 資料從指定儲存格B10開始輸出報名名單
    start_row = 10
    start_column = 2  # 欄位 B
    for index, row in df2.iterrows():
        for col_num, value in enumerate(row, start=start_column):
            sheet.cell(row=start_row + index, column=col_num, value=value)
    
    # 輸出報名表表頭資料至報名表照片sheet
    sheet = workbook['報名表照片']
    sheet['C2'].value = df1['game_id'].values[0]+'報名表(照片)'
    sheet['B3'].value = df1['報名單位'].values[0]
    sheet['E3'].value = df1['聯絡人'].values[0]
    sheet['B4'].value = df1['電話'].values[0]
    sheet['E4'].value = df1['LINE_ID'].values[0]
    sheet['B5'].value = df1['電子郵件'].values[0]
    sheet['B6'].value = df1['參賽組別'].values[0]
    # 輸出報名編號至儲存格
    #names = df2['姓名'].tolist()
    '''
    names = [row['姓名'] for row in data]
    num_names = len(names)
    start_row = 7
    start_column = 1

    for i in range(num_names):
        row = start_row + (i // 5) * 2
        column = start_column + i % 5
        sheet.cell(row=row, column=column, value=f"{i+1}.姓名：{names[i]}")
    '''
    # 将st_data按规则输出到Excel表格, 從A8開始輸出相片, A9開始輸出姓名
    num_photos = 0
    start_row = 8
    start_column = 1 #欄位 A

    for row in data:
        team_num = row['team_num']
        name = row['姓名']
        photo_data = row['st_data']
        
        if photo_data: # 若相片存在
            # 将二进制图片数据转换为PIL图片对象
            image = Image.open(BytesIO(photo_data))
            
            # 暫存相片檔案至 EXPORT_FOLDER 
            folder_path = os.path.join(app.root_path, app.config['EXPORT_FOLDER'])
            file_path = os.path.join(folder_path,f"{team_num}_{name}.jpg")
            image.save(file_path)
            
            # 创建Excel图片对象
            img = drawing.image.Image(file_path)
            # 调整图片的大小和位置
            img.width = 130
            img.height = 110                                   
            img_cell = sheet.cell(row=start_row, column=start_column)
            sheet.add_image(img, img_cell.coordinate)

        # 将姓名按规则输出到Excel表格
        name_cell = sheet.cell(row=start_row + 1, column=start_column)
        name_cell.value = f"{num_photos+1}.姓名：{name}"
        # 控制每列最多輸出5張圖片及最多輸出5個姓名
        num_photos += 1
        if num_photos % 5 == 0:
            start_row += 2
            start_column = 1
        else:
            start_column += 1
        

    #for row in dataframe_to_rows(df2, index=False, header=False):
    #    sheet.append(row)
    '''# 採用落地存檔,而將存檔直接輸出download
    folder_path = os.path.join(app.root_path, app.config['EXPORT_FOLDER'])
    file_path = os.path.join(folder_path, f"{df1['報名單位'].values[0]}_{df1['參賽組別'].values[0]}.xlsx")
    workbook.save(file_path)
    return send_file(file_path, as_attachment=True)
    '''
    # 採用不落地存檔,而將記憶體直接寫出至excel 
    # 建立一個記憶體緩衝區 (in-memory buffer)
    buffer = BytesIO()
    # 將 Workbook 寫入緩衝區
    workbook.save(buffer)
    buffer.seek(0)  # 將緩衝區指標移回起始位置
    
    # 刪除相片暫存檔案
    ToDel_folder_path = os.path.join(app.root_path,app.config['EXPORT_FOLDER'])
    # 遍历文件夹中的所有文件
    for filename in os.listdir(ToDel_folder_path):
        file_path = os.path.join(ToDel_folder_path, filename)       
        # 检查文件是否为JPEG格式
        if filename.startswith(str(team_num)) and filename.lower().endswith('.jpg'):
            # 删除文件
            os.remove(file_path)
    
    return send_file(buffer, as_attachment=True, download_name=f"{df1['報名單位'].values[0]}_{df1['參賽組別'].values[0]}.xlsx",
    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# Download 報名球員大頭照, 下載檔案格式為zip
@app.route('/download_photo/<int:team_id>', methods=['GET'])
@login_required
@roles_accepted('admin', 'gamemanager', 'user')
def download_photo(team_id):
    sql1 = f"SELECT team_name FROM team WHERE team_id={team_id} "
    team_name = engine.execute(sql1).fetchone()[0]

    sql2 = f'''SELECT st_data,student_name
        FROM registration WHERE team_num={team_id} '''
    rows = engine.execute(sql2)
    
    # Create a BytesIO object to store the zip file
    zip_data = BytesIO()

    # Create a ZipFile object to write the photo
    with zipfile.ZipFile(zip_data, 'w') as zip_file:
        # Iterate through the rows and add each photo to the zip file
        for i, row in enumerate(rows):
            if row['st_data']: # if photo data existed
                image_data = row['st_data']  
                file_name = f"{team_name}_{row['student_name']}.jpg"  # Set a file name for the photo
                zip_file.writestr(file_name, image_data)

    # Set the appropriate headers for the zip file response
    response = make_response(zip_data.getvalue())
    response.headers['Content-Disposition'] = f'''attachment; filename="{quote(team_name)}_photos.zip" '''
    response.headers['Content-Type'] = 'application/zip'

    return response


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=8000, debug=True)

